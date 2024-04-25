import openpyxl
import xlsxwriter


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


def get_data_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    column_count = sheet.max_column

    for r in range(2, row_count + 1):
        row_list = []
        for c in range(1, column_count + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list


def generate_excel(workbook_name: str, worksheet_name: str, headers_list: list, data: list):
    workbook = xlsxwriter.workbook.Workbook(workbook_name)

    worksheet = workbook.add_worksheet(worksheet_name)

    for index, header in enumerate(headers_list):
        worksheet.write(0, index, str(header).capitalize())

    for index1, entry in enumerate(data):
        for index2, header in enumerate(headers_list):
            worksheet.write(index1+1, index2, entry[header])

    workbook.close()

